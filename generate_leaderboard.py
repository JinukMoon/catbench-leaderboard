#!/usr/bin/env python3
"""
CatBench Leaderboard Generator
This script aggregates benchmark results from multiple datasets and generates
an interactive leaderboard website for React deployment.
"""

import json
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings
import openpyxl
warnings.filterwarnings('ignore')


class LeaderboardGenerator:
    """Generate leaderboard data from CatBench benchmark results."""

    def __init__(self, results_dir='results/cathub', output_dir='public'):
        self.results_dir = Path(results_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        # Define key metrics to track
        self.key_metrics = [
            'MAE_total (eV)',
            'MAE_normal (eV)',
            'Normal rate (%)',
            'ADwT (%)',
            'Time_per_step (s)'
        ]

        # Store all aggregated data
        self.leaderboard_data = {}
        self.dataset_info = {}
        self.mlip_info = {}

    def collect_benchmark_data(self):
        """Collect all benchmark data from Excel files."""
        print("📊 Collecting benchmark data...")

        # Find all Excel files
        excel_files = list(self.results_dir.glob('*/*_Benchmarking_Analysis.xlsx'))
        print(f"Found {len(excel_files)} benchmark files")

        for excel_file in excel_files:
            dataset_name = excel_file.parent.name
            print(f"  Processing {dataset_name}...")

            try:
                # Read MLIP_Data sheet
                df = pd.read_excel(excel_file, sheet_name='MLIP_Data')

                # Clean the dataframe (remove NaN rows)
                df = df.dropna(subset=['MLIP_name'])

                # Store dataset info
                if dataset_name not in self.dataset_info:
                    # Calculate total structures from all MLIPs
                    num_total = 0
                    if 'Num_total' in df.columns and not df.empty:
                        # Sum up Num_total from all MLIPs (they should be the same, but take max to be safe)
                        num_total = int(df['Num_total'].max()) if not df['Num_total'].isna().all() else 0
                    
                    self.dataset_info[dataset_name] = {
                        'name': dataset_name,
                        'num_structures': num_total,
                        'file_path': str(excel_file.absolute())  # Store absolute path
                    }

                # Process each MLIP
                for _, row in df.iterrows():
                    mlip_name = row['MLIP_name']

                    # Initialize MLIP data structure
                    if mlip_name not in self.leaderboard_data:
                        self.leaderboard_data[mlip_name] = {
                            'datasets': {},
                            'average_metrics': {}
                        }

                    # Store metrics for this dataset
                    metrics = {}
                    for metric in self.key_metrics:
                        # Handle different column name formats
                        metric_key = metric
                        if metric_key not in row:
                            # Try alternative names
                            alt_names = {
                                'MAE_total (eV)': ['MAE_total (eV)', 'MAE_total', 'MAE_total(eV)'],
                                'MAE_normal (eV)': ['MAE_normal (eV)', 'MAE_normal', 'MAE_normal(eV)'],
                                'Normal rate (%)': ['Normal rate (%)', 'Normal rate', 'Normal rate(%)', 'Normal (%)'],
                                'ADwT (%)': ['ADwT (%)', 'ADwT', 'ADwT(%)'],
                                'Time_per_step (s)': ['Time_per_step (s)', 'Time_per_step', 'Time/step (s)', 'Time/step']
                            }
                            for alt in alt_names.get(metric_key, []):
                                if alt in row and pd.notna(row[alt]):
                                    metric_key = alt
                                    break
                        
                        if metric_key in row and pd.notna(row[metric_key]):
                            metrics[metric] = float(row[metric_key])

                    self.leaderboard_data[mlip_name]['datasets'][dataset_name] = metrics

            except Exception as e:
                print(f"    ⚠️ Error processing {dataset_name}: {e}")
                import traceback
                traceback.print_exc()
                continue

    def calculate_aggregate_metrics(self):
        """Calculate aggregate metrics across all datasets for each MLIP."""
        print("\n📈 Calculating aggregate metrics...")

        for mlip_name, mlip_data in self.leaderboard_data.items():
            datasets = mlip_data['datasets']

            if not datasets:
                continue

            # Calculate averages across datasets
            avg_metrics = {}
            for metric in self.key_metrics:
                values = []
                weights = []  # For weighted average based on dataset size
                
                for dataset_name, dataset_metrics in datasets.items():
                    if metric in dataset_metrics:
                        values.append(dataset_metrics[metric])
                        # Use dataset size as weight for weighted average
                        dataset_size = self.dataset_info.get(dataset_name, {}).get('num_structures', 1)
                        weights.append(dataset_size)

                if values:
                    # For MAE metrics, use weighted average (larger datasets have more weight)
                    if 'MAE' in metric:
                        if sum(weights) > 0:
                            weighted_mean = sum(v * w for v, w in zip(values, weights)) / sum(weights)
                        else:
                            weighted_mean = float(np.mean(values))
                        avg_metrics[metric] = {
                            'mean': weighted_mean,
                            'std': float(np.std(values)),
                            'min': float(np.min(values)),
                            'max': float(np.max(values)),
                            'count': len(values)
                        }
                    else:
                        # For other metrics, use simple average
                        avg_metrics[metric] = {
                            'mean': float(np.mean(values)),
                            'std': float(np.std(values)),
                            'min': float(np.min(values)),
                            'max': float(np.max(values)),
                            'count': len(values)
                        }

            mlip_data['average_metrics'] = avg_metrics

            # Calculate overall score (lower is better for MAE, higher for success rate)
            score_components = []

            # MAE contribution (normalized, lower is better)
            if 'MAE_total (eV)' in avg_metrics:
                mae_score = 1.0 / (1.0 + avg_metrics['MAE_total (eV)']['mean'])
                score_components.append(mae_score * 0.4)  # 40% weight

            # Success rate contribution (higher is better)
            if 'Normal rate (%)' in avg_metrics:
                success_score = avg_metrics['Normal rate (%)']['mean'] / 100.0
                score_components.append(success_score * 0.4)  # 40% weight

            # Speed contribution (lower is better, normalized)
            if 'Time_per_step (s)' in avg_metrics:
                speed_score = 1.0 / (1.0 + avg_metrics['Time_per_step (s)']['mean'])
                score_components.append(speed_score * 0.2)  # 20% weight

            mlip_data['overall_score'] = sum(score_components) if score_components else 0
            mlip_data['num_datasets'] = len(datasets)

    def generate_rankings(self):
        """Generate rankings for different metrics."""
        print("\n🏆 Generating rankings...")

        rankings = {
            'overall': [],
            'accuracy': [],
            'success_rate': [],
            'speed': [],
            'coverage': []
        }

        # Overall ranking (by composite score)
        for mlip_name, mlip_data in self.leaderboard_data.items():
            if 'overall_score' in mlip_data:
                rankings['overall'].append({
                    'mlip': mlip_name,
                    'score': mlip_data['overall_score'],
                    'num_datasets': mlip_data['num_datasets']
                })
        rankings['overall'].sort(key=lambda x: x['score'], reverse=True)

        # Accuracy ranking (by MAE)
        for mlip_name, mlip_data in self.leaderboard_data.items():
            if 'average_metrics' in mlip_data and 'MAE_total (eV)' in mlip_data['average_metrics']:
                rankings['accuracy'].append({
                    'mlip': mlip_name,
                    'mae': mlip_data['average_metrics']['MAE_total (eV)']['mean'],
                    'std': mlip_data['average_metrics']['MAE_total (eV)']['std']
                })
        rankings['accuracy'].sort(key=lambda x: x['mae'])

        # Success rate ranking
        for mlip_name, mlip_data in self.leaderboard_data.items():
            if 'average_metrics' in mlip_data and 'Normal rate (%)' in mlip_data['average_metrics']:
                rankings['success_rate'].append({
                    'mlip': mlip_name,
                    'rate': mlip_data['average_metrics']['Normal rate (%)']['mean'],
                    'std': mlip_data['average_metrics']['Normal rate (%)']['std']
                })
        rankings['success_rate'].sort(key=lambda x: x['rate'], reverse=True)

        # Speed ranking
        for mlip_name, mlip_data in self.leaderboard_data.items():
            if 'average_metrics' in mlip_data and 'Time_per_step (s)' in mlip_data['average_metrics']:
                rankings['speed'].append({
                    'mlip': mlip_name,
                    'time': mlip_data['average_metrics']['Time_per_step (s)']['mean'],
                    'std': mlip_data['average_metrics']['Time_per_step (s)']['std']
                })
        rankings['speed'].sort(key=lambda x: x['time'])

        # Coverage ranking (number of datasets tested)
        for mlip_name, mlip_data in self.leaderboard_data.items():
            rankings['coverage'].append({
                'mlip': mlip_name,
                'count': mlip_data['num_datasets']
            })
        rankings['coverage'].sort(key=lambda x: x['count'], reverse=True)

        return rankings

    def extract_excel_data(self):
        """Extract full Excel data for each dataset, including adsorbate breakdown."""
        print("\n📋 Extracting Excel data for dataset views...")
        
        excel_data = {}
        adsorbate_breakdown = {}
        
        for dataset_name, dataset_info in self.dataset_info.items():
            excel_file_path = Path(dataset_info['file_path'])
            
            # Ensure absolute path
            if not excel_file_path.is_absolute():
                # Try relative to current directory first
                if (Path.cwd() / excel_file_path).exists():
                    excel_file_path = Path.cwd() / excel_file_path
                # Then try relative to results_dir
                elif (self.results_dir / dataset_name / f"{dataset_name}_Benchmarking_Analysis.xlsx").exists():
                    excel_file_path = self.results_dir / dataset_name / f"{dataset_name}_Benchmarking_Analysis.xlsx"
                else:
                    print(f"    ⚠️ Excel file not found: {excel_file_path}")
                    continue
            
            if not excel_file_path.exists():
                print(f"    ⚠️ Excel file not found: {excel_file_path}")
                continue
            
            try:
                # Read Excel file with openpyxl to handle merged cells
                wb = openpyxl.load_workbook(excel_file_path, data_only=True)
                dataset_sheets = {}
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Skip summary sheets
                    if sheet_name.lower() in ['mlip_data', 'anomaly_analysis', 'summary', 'anomaly']:
                        continue
                    
                    # Process merged cells to create proper headers
                    merged_ranges = list(ws.merged_cells.ranges)
                    
                    # Read all rows
                    all_rows = []
                    max_row = ws.max_row
                    max_col = ws.max_column
                    
                    for row_idx in range(1, max_row + 1):
                        row_data = []
                        for col_idx in range(1, max_col + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            value = cell.value
                            
                            # Check if cell is part of a merged range
                            if value is None:
                                for merged_range in merged_ranges:
                                    if cell.coordinate in merged_range:
                                        # Get the top-left cell value
                                        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                        value = top_left.value
                                        break
                            
                            # Format numeric values
                            if isinstance(value, (int, float)) and not isinstance(value, bool):
                                if abs(value) < 1e-10:  # Very small numbers
                                    value = 0.0
                                else:
                                    value = float(value)
                            
                            row_data.append(value if value is not None else '')
                        all_rows.append(row_data)
                    
                    # Extract headers (first row or first 2 rows for merged headers)
                    if len(all_rows) >= 1:
                        # Try to detect if first row is header
                        header_row = all_rows[0]
                        
                        # Check if second row might be part of header
                        if len(all_rows) >= 2:
                            second_row = all_rows[1]
                            # If second row has mostly text/headers, combine
                            if any(isinstance(v, str) and v.strip() for v in second_row[:5]):
                                headers = []
                                for i in range(len(header_row)):
                                    h1 = str(header_row[i]) if header_row[i] else ''
                                    h2 = str(second_row[i]) if i < len(second_row) and second_row[i] else ''
                                    
                                    # Special handling for Anomaly count columns
                                    if h1 == 'Anomaly count' and h2:
                                        # Create separate columns for each detection scheme
                                        headers.append(f"Anomaly count - {h2}")
                                    elif h1 == 'Anomaly count' and not h2:
                                        headers.append('Anomaly count - total')
                                    elif h1 and h2 and h1 != h2:
                                        headers.append(f"{h1} - {h2}")
                                    elif h1:
                                        headers.append(h1)
                                    elif h2:
                                        headers.append(h2)
                                    else:
                                        headers.append(f"Column {i+1}")
                                data_rows = all_rows[2:] if len(all_rows) > 2 else []
                            else:
                                headers = [str(h) if h else f"Column {i+1}" for i, h in enumerate(header_row)]
                                data_rows = all_rows[1:] if len(all_rows) > 1 else []
                        else:
                            headers = [str(h) if h else f"Column {i+1}" for i, h in enumerate(header_row)]
                            data_rows = []
                        
                        # Store sheet data (all sheets including MLIP-specific ones)
                        dataset_sheets[sheet_name] = {
                            'columns': headers,
                            'data': data_rows
                        }
                        
                        # Extract adsorbate breakdown for MLIP-specific sheets
                        # MLIP sheet names are typically the MLIP name itself
                        # Check if this sheet name matches any known MLIP name
                        mlip_name = sheet_name
                        # Store in adsorbate_breakdown for backward compatibility
                        # (use first dataset's data for each MLIP, or merge if needed)
                        if mlip_name not in adsorbate_breakdown:
                            adsorbate_breakdown[mlip_name] = {
                                'columns': headers,
                                'data': data_rows
                            }
                
                excel_data[dataset_name] = dataset_sheets
                print(f"    ✅ Extracted data from {dataset_name} ({len(wb.sheetnames)} sheets)")
                
            except Exception as e:
                print(f"    ⚠️ Error extracting Excel data for {dataset_name}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        return excel_data, adsorbate_breakdown

    def save_json_data(self):
        """Save all data as JSON for the web interface."""
        print("\n💾 Saving JSON data...")

        # Generate rankings
        rankings = self.generate_rankings()
        
        # Extract Excel data for dataset detail views and adsorbate breakdown
        excel_data, adsorbate_breakdown = self.extract_excel_data()

        # Prepare final data structure
        final_data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'num_mlips': len(self.leaderboard_data),
                'num_datasets': len(self.dataset_info),
                'metrics': self.key_metrics
            },
            'mlips': self.leaderboard_data,
            'datasets': self.dataset_info,
            'rankings': rankings,
            'excel_data': excel_data,
            'adsorbate_breakdown': adsorbate_breakdown
        }

        # Save to JSON file
        json_path = self.output_dir / 'leaderboard_data.json'
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(final_data, f, indent=2, ensure_ascii=False)

        print(f"  ✅ Saved data to {json_path}")
        print(f"     - {len(self.leaderboard_data)} MLIPs")
        print(f"     - {len(self.dataset_info)} datasets")
        print(f"     - {len(adsorbate_breakdown)} MLIPs with adsorbate breakdown")

        return final_data

    def generate_summary_report(self, data):
        """Generate a text summary report."""
        print("\n📝 Generating summary report...")

        report = []
        report.append("=" * 80)
        report.append("CATBENCH LEADERBOARD SUMMARY")
        report.append("=" * 80)
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append(f"Total MLIPs evaluated: {data['metadata']['num_mlips']}")
        report.append(f"Total datasets: {data['metadata']['num_datasets']}")
        report.append("")

        # Top performers
        report.append("TOP PERFORMERS BY CATEGORY")
        report.append("-" * 40)

        if data['rankings']['overall']:
            report.append("\n🏆 Overall Score:")
            for i, item in enumerate(data['rankings']['overall'][:5], 1):
                report.append(f"  {i}. {item['mlip']}: {item['score']:.3f} ({item['num_datasets']} datasets)")

        if data['rankings']['accuracy']:
            report.append("\n🎯 Best Accuracy (MAE):")
            for i, item in enumerate(data['rankings']['accuracy'][:5], 1):
                report.append(f"  {i}. {item['mlip']}: {item['mae']:.3f} ± {item['std']:.3f} eV")

        if data['rankings']['success_rate']:
            report.append("\n✅ Highest Success Rate:")
            for i, item in enumerate(data['rankings']['success_rate'][:5], 1):
                report.append(f"  {i}. {item['mlip']}: {item['rate']:.1f} ± {item['std']:.1f}%")

        if data['rankings']['speed']:
            report.append("\n⚡ Fastest Models:")
            for i, item in enumerate(data['rankings']['speed'][:5], 1):
                report.append(f"  {i}. {item['mlip']}: {item['time']:.4f} ± {item['std']:.4f} s/step")

        if data['rankings']['coverage']:
            report.append("\n📊 Best Coverage:")
            for i, item in enumerate(data['rankings']['coverage'][:5], 1):
                report.append(f"  {i}. {item['mlip']}: {item['count']} datasets")

        report.append("\n" + "=" * 80)

        # Save report
        report_path = self.output_dir / 'summary_report.txt'
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(report))

        print(f"  ✅ Saved report to {report_path}")

        # Also print to console
        print("\n" + '\n'.join(report))

    def run(self):
        """Run the complete leaderboard generation process."""
        print("\n🚀 Starting CatBench Leaderboard Generation")
        print("=" * 50)

        # Collect data
        self.collect_benchmark_data()

        # Calculate aggregates
        self.calculate_aggregate_metrics()

        # Save JSON data
        data = self.save_json_data()

        # Generate summary report
        self.generate_summary_report(data)

        print("\n✨ Leaderboard generation complete!")
        print(f"   Output directory: {self.output_dir}")

        return data


def main():
    """Main entry point."""
    generator = LeaderboardGenerator(
        results_dir='results/cathub',
        output_dir='public'
    )
    generator.run()


if __name__ == "__main__":
    main()

